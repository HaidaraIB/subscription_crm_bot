from __future__ import annotations

import os
import tempfile
from dataclasses import dataclass, field
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import models
from common.lang_dicts import TEXTS
from common.subscription_utils import (
    compute_end_date,
    format_date,
    normalize_phone,
    parse_date,
)

# (TEXTS key, field name for import)
CUSTOMER_COLUMNS: list[tuple[str, str]] = [
    ("excel_customer_id", "id"),
    ("excel_customer_name", "name"),
    ("excel_phone", "phone"),
    ("excel_service_username", "service_username"),
    ("excel_service_password", "service_password"),
    ("excel_subscription_type", "subscription_type"),
    ("excel_duration_days", "duration_days"),
    ("excel_start_date", "start_date"),
    ("excel_end_date", "end_date"),
    ("excel_telegram_user_id", "telegram_user_id"),
    ("excel_notes", "notes"),
    ("excel_customer_status", "status"),
]

IMPORT_FIELDS = {f for _, f in CUSTOMER_COLUMNS if f not in ("id", "status")}


def customer_headers(lang: models.Language) -> list[str]:
    return [TEXTS[lang][key] for key, _ in CUSTOMER_COLUMNS]


def _header_aliases() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for text_key, field in CUSTOMER_COLUMNS:
        for lang in (models.Language.ARABIC, models.Language.ENGLISH):
            label = TEXTS[lang][text_key].strip().lower()
            aliases[label] = field
        aliases[field] = field
        aliases[text_key.replace("excel_", "")] = field
    return aliases


def _style_header_row(ws) -> None:
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(ws, widths: list[int]) -> None:
    from openpyxl.utils import get_column_letter

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def export_customers_workbook(lang: models.Language, session) -> str:
    """Write all customers to a temp .xlsx path and return the path."""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    path = tmp.name

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(customer_headers(lang))
    _style_header_row(ws)

    today = date.today()
    customers = session.query(models.Customer).order_by(models.Customer.id.asc()).all()
    for c in customers:
        status_key = (
            "subs_status_active" if c.end_date >= today else "subs_status_expired"
        )
        ws.append(
            [
                c.id,
                c.name or "",
                c.phone,
                c.service_username,
                c.service_password,
                c.subscription_type,
                c.duration_days,
                format_date(c.start_date),
                format_date(c.end_date),
                c.telegram_user_id if c.telegram_user_id is not None else "",
                c.notes or "",
                TEXTS[lang][status_key],
            ]
        )

    _autosize_columns(
        ws,
        [8, 22, 16, 18, 18, 16, 12, 14, 14, 16, 28, 12],
    )
    wb.save(path)
    return path


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_to_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return parse_date(_cell_to_str(value))


def _cell_to_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _cell_to_telegram_id(value) -> int | None:
    text = _cell_to_str(value)
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _map_headers(header_row: list) -> dict[int, str]:
    aliases = _header_aliases()
    mapping: dict[int, str] = {}
    for col_idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = _cell_to_str(cell).lower()
        field = aliases.get(key)
        if field and field in IMPORT_FIELDS | {"id"}:
            mapping[col_idx] = field
    return mapping


@dataclass
class ImportRowError:
    row: int
    message: str


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[ImportRowError] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors and (self.created or self.updated)


def _parse_row(
    row_values: list,
    col_map: dict[int, str],
    row_num: int,
    lang: models.Language,
) -> tuple[dict | None, str | None]:
    data: dict = {}
    for col_idx, field in col_map.items():
        if col_idx < len(row_values):
            data[field] = row_values[col_idx]

    phone_raw = _cell_to_str(data.get("phone"))
    if not phone_raw:
        return None, TEXTS[lang]["subs_import_missing_phone"]
    phone = normalize_phone(phone_raw)
    if not phone:
        return None, TEXTS[lang]["subs_invalid_phone"]

    username = _cell_to_str(data.get("service_username"))
    password = _cell_to_str(data.get("service_password"))
    if not username:
        return None, TEXTS[lang]["subs_import_missing_username"]
    if not password:
        return None, TEXTS[lang]["subs_import_missing_password"]

    duration = _cell_to_int(data.get("duration_days"))
    if duration is None or duration <= 0:
        return None, TEXTS[lang]["subs_invalid_duration"]

    start = _cell_to_date(data.get("start_date"))
    if not start:
        return None, TEXTS[lang]["subs_invalid_date"]

    end = _cell_to_date(data.get("end_date"))
    if not end:
        end = compute_end_date(start, duration)

    sub_type = _cell_to_str(data.get("subscription_type"))

    customer_id = _cell_to_int(data.get("id"))
    name = _cell_to_str(data.get("name")) or None
    notes = _cell_to_str(data.get("notes")) or None
    telegram_id = _cell_to_telegram_id(data.get("telegram_user_id"))

    return {
        "id": customer_id,
        "name": name,
        "phone": phone,
        "service_username": username,
        "service_password": password,
        "subscription_type": sub_type,
        "duration_days": duration,
        "start_date": start,
        "end_date": end,
        "notes": notes,
        "telegram_user_id": telegram_id,
    }, None


def _apply_customer_fields(customer: models.Customer, parsed: dict) -> None:
    customer.name = parsed["name"]
    customer.phone = parsed["phone"]
    customer.service_username = parsed["service_username"]
    customer.service_password = parsed["service_password"]
    customer.subscription_type = parsed["subscription_type"]
    customer.duration_days = parsed["duration_days"]
    customer.start_date = parsed["start_date"]
    customer.end_date = parsed["end_date"]
    customer.notes = parsed["notes"]
    customer.telegram_user_id = parsed["telegram_user_id"]


def import_customers_workbook(
    path: str, lang: models.Language, session
) -> ImportResult:
    result = ImportResult()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        result.errors.append(ImportRowError(0, TEXTS[lang]["subs_import_no_rows"]))
        return result

    col_map = _map_headers(list(rows[0]))
    required = {
        "phone",
        "service_username",
        "service_password",
        "duration_days",
        "start_date",
    }
    if not required.issubset(set(col_map.values())):
        result.errors.append(ImportRowError(1, TEXTS[lang]["subs_import_wrong_format"]))
        return result

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        parsed, err = _parse_row(list(row), col_map, row_num, lang)
        if err:
            result.errors.append(ImportRowError(row_num, err))
            continue
        if parsed is None:
            result.skipped += 1
            continue

        customer = None
        if parsed["id"]:
            customer = session.get(models.Customer, parsed["id"])

        if customer:
            _apply_customer_fields(customer, parsed)
            result.updated += 1
        else:
            customer = models.Customer(
                name=parsed["name"],
                phone=parsed["phone"],
                service_username=parsed["service_username"],
                service_password=parsed["service_password"],
                subscription_type=parsed["subscription_type"],
                duration_days=parsed["duration_days"],
                start_date=parsed["start_date"],
                end_date=parsed["end_date"],
                notes=parsed["notes"],
                telegram_user_id=parsed["telegram_user_id"],
            )
            session.add(customer)
            result.created += 1

    return result


def safe_unlink(path: str | None) -> None:
    if path and os.path.exists(path):
        os.unlink(path)
